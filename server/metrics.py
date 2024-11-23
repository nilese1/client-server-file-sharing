import ntplib
import pandas as pd
from pathlib import Path
import openpyxl


class Metrics:
    def __init__(self, client_ip, excelPath: Path):
        self.ntpClient = ntplib.NTPClient()
        self.client_ip = client_ip
        self.metricsDf = pd.DataFrame(
            columns=["ip", "operationType", "elapsedTime", "transferRate", "bytes"]
        )
        self.excelPath = excelPath

        # create excel file to permanently store metrics
        if not excelPath.exists():
            wb = openpyxl.Workbook()
            wb.save(excelPath)

    # store in dataframe. probably write to external file later (excel or csv)
    def calculateMetrics(self, operationType, ntpStart, ntpEnd, bytes_transferred):
        ipstring = self.client_ip[0] + ":" + str(self.client_ip[1])
        elapsedTime = -1 if (ntpStart == -1 or ntpEnd == -1) else ntpEnd - ntpStart
        transferRate = (
            -1 if (elapsedTime == -1) else (bytes_transferred / 1024) / elapsedTime
        )  # kbps
 
        self.metricsDf.loc[len(self.metricsDf.index)] = [
            ipstring,
            operationType,
            elapsedTime,
            transferRate,
            bytes_transferred,
        ]

    def processMetrics(self):
        try:
            sheetName = "metrics"
            workbook = openpyxl.load_workbook(self.excelPath)

            if not sheetName in workbook.sheetnames:
                # if sheet is blank, insert rows as normal
                with pd.ExcelWriter(
                    self.excelPath,
                    engine="openpyxl",
                    mode="a",
                    if_sheet_exists="replace",
                ) as writer:
                    self.metricsDf.to_excel(writer, sheet_name=sheetName, index=False)
            else:
                # get sheet data and append to it. then insert combined data as normal
                sheetDf = pd.read_excel(self.excelPath, sheet_name=sheetName)
                combinedDf = pd.concat([sheetDf, self.metricsDf], ignore_index=True)
                with pd.ExcelWriter(
                    self.excelPath,
                    engine="openpyxl",
                    mode="a",
                    if_sheet_exists="replace",
                ) as writer:
                    combinedDf.to_excel(writer, sheet_name=sheetName, index=False)

            print(f"Metrics saved to {self.excelPath}")
        except Exception as e:
            print(f"Error processing metrics {e}")

    def getNTPTime(self):
        try:
            ntp_res = self.ntpClient.request("0.pool.ntp.org", version=3)
            return ntp_res.tx_time
        except Exception as e:
            print(f"Error getting NTP time, returning -1")
            return -1
